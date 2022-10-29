from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http.response import HttpResponse
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Job, Employee
from openpyxl import Workbook

import datetime
import calendar
import mimetypes
import os





def loginmanager(request):
    if request.method == "POST":
        uname = request.POST['uname']
        pwd = request.POST['pwd']
        user = authenticate(request, username=uname, password=pwd)
        if user:
            if user.is_staff == 1:
                login(request, user)
                return redirect('ind')
        # messages.error(request, 'Incorrect username or Password', extra_tags='text-danger')
    return render(request, 'login.html')


# logout function
def logoutUser(request):
    logout(request)
    return redirect('login')

# dashboard(index)


@login_required(login_url='login')
def index(request):

    return render(request, 'index.html')


# managentreport
@login_required(login_url='login')
def manage(request):
    data = request.GET
    emps = Employee.objects.order_by('name')
    now = datetime.datetime.now()
    dates = f"{now.day} - {calendar.month_name[now.month]} - {now.year} - {now.hour} : {now.minute}"
    date = f"{now.year}:{now.month}:{now.day}"
    sdate = f'{now.year}{now.month}{now.day}'

    if request.GET == True:
        empl = data.get('empl')
        jobs = Job.objects.filter(fdate=None, user_id=empl,)
        context = {'dates': dates, 'emps': emps, 'jobs': jobs}
        return render(request, 'dashboard.html', context)

    if request.method == 'POST':

        data = request.POST
        p = data['add']
        if p == 'add':
            fro = data['from']
            if fro == '':
                pass
            else:
                fro =fro.replace("-", ":")
                date=fro
            job = data['jobadd']
            task = data['task']
            refer = data['refer']
            user = data['user']
            Job.objects.create(
                jobname=job,
                task=task,
                ref=refer,
                sdate=date,
                status='Not completed',
                user_id=user,
                day=sdate,
            )
            # filter employee details
            jobs = Job.objects.filter(user_id=user, status='Not completed')
            nub = len(jobs)
            nocmpj = Job.objects.filter(user_id=user)
            bewu = len(nocmpj)
            comp = bewu-nub
            # update table
            emps = Employee.objects.filter(id=user)
            em = Employee.objects.get(id=user)
            em.progress = nub
            em.task = bewu
            em.completed = comp
            em.save()

            return redirect('manage')
    context = {'dates': dates, 'emps': emps, }
    return render(request, 'dashboard.html', context)


@login_required(login_url='login')
def prev(request):
    # date time
    now = datetime.datetime.now()
    dates = f"{now.day} - {calendar.month_name[now.month]} - {now.year} - {now.hour} : {now.minute}"
    sdate = f'{now.year}{now.month}{now.day}'
    # get
    data = request.GET
    empl = data.get('empl')
    # filter employee details
    jobs = Job.objects.filter(user_id=empl, status='Not completed')
    nub = len(jobs)
    nocmpj = Job.objects.filter(user_id=empl)
    bewu = len(nocmpj)
    comp = bewu-nub
    # update table
    emps = Employee.objects.filter(id=empl)
    em = Employee.objects.get(id=empl)
    em.progress = nub
    em.task = bewu
    em.completed = comp
    em.save()
    jobtds = Job.objects.filter(
        user_id=empl, day=sdate, status='Not completed')

    # find which button clicked
    if request.method == 'POST':
        button = request.POST.get('old')
        buttonnew = request.POST.get('new')
        if button == 'true':
            # get form value
            data = request.POST
            job = data['job']
            task = data['task']
            remark = data['remark']
            uid = data['user']
            status = data['status']
            if status == "Completed":
                fdate = dates
            else:
                fdate = None
            jobs = Job.objects.get(id=uid)
            jobs.jobname = job
            jobs.task = task
            jobs.remark = remark
            jobs.status = status
            jobs.fdate = fdate
            jobs.save()
            # refresh  employee status
            jobs = Job.objects.filter(user_id=empl, status='Not completed')
            nub = len(jobs)
            nocmpj = Job.objects.filter(user_id=empl)
            bewu = len(nocmpj)
            comp = bewu-nub
            # update
            em = Employee.objects.get(id=empl)
            em.progress = nub
            em.task = bewu
            em.completed = comp
            em.save()
        #  if new task button clicked
        elif buttonnew == 'new':
            data = request.POST
            job = data['job']
            task = data['task']
            remark = data['remark']
            uid = data['user']
            status = data['status']
            if status == "Completed":
                fdate = dates
            else:
                fdate = None
            jobs = Job.objects.get(id=uid)
            jobs.jobname = job
            jobs.task = task
            jobs.remark = remark
            jobs.status = status
            jobs.fdate = dates
            jobs.save()

            jobs = Job.objects.filter(user_id=empl, status='Not completed')
            nub = len(jobs)
            nocmpj = Job.objects.filter(user_id=empl)
            bewu = len(nocmpj)
            comp = bewu-nub
            em = Employee.objects.get(id=empl)
            em.progress = nub
            em.task = bewu
            em.completed = comp
            em.save()
        return redirect('manage')

    context = {'jobs': jobs, 'emps': emps,
               'dates': dates, 'jobtds': jobtds, 'sdate': sdate}
    return render(request, 'prev.html', context)


@login_required(login_url='login')
def addemp(request):
    if request.method == 'POST':
        data = request.POST
        emp = data['empname']
        detail = data['details']
        phone = data['phone']
        email = data['email']
        Employee.objects.create(
            name=emp,
            email=email,
            phone=phone,
            details=detail
        )

    return render(request, 'addemployee.html')


# take report of all employee
@login_required(login_url='login')
def report(request):
    emps = Employee.objects.order_by('name')
    # form submit for take report with date range
    if request.method == 'POST':
        # get from and to value
        data = request.POST
        fro = data['from']
        to = data['to']
        # for report of one person and view all jobs and details
        try:
            uid = data['user']
            if fro == '':
                jobs = Job.objects.get(id=1)
                fro = jobs.day
                fro = int(fro.replace("-", ""))
            else:
                fro = int(fro.replace("-", ""))
            if to == '':
                now = datetime.datetime.now()
                sdate = f'{now.year}{now.month}{now.day}'
                to = int(sdate.replace("-", ""))
            else:
                to = int(to.replace("-", ""))
            # get number of jobs
            lenj = 0
            nub = 0
            jsj = list()
            jsjn=list()
            jsjc=list()
            for i in range(fro, to):
                jobs = Job.objects.filter(day=i, user_id=uid)  # total
                for job in jobs:
                    jsj.append(job.id)
                lenj = lenj+len(jobs)
                jobss = Job.objects.filter(
                    user_id=uid, day=i, status='Not completed')  # not completed
                for job in jobss:
                    jsjn.append(job.id)
                jobst = Job.objects.filter(
                    user_id=uid, day=i, status='Completed')  # completed
                for job in jobst:
                    jsjc.append(job.id)
                nub = nub+len(jobss)
                comp = lenj-nub
                # refresh employee status
            # for excel report of employee
            emps = Employee.objects.filter(id=uid)
            jobs = Job.objects.filter(user_id=uid)
            emp = Employee.objects.get(id=uid)
            # Excel work book create
            wb=Workbook()
            wb.create_sheet("sheet_one")
            ws1=wb['sheet_one']
            ws1['A1']='Employee Name'
            ws1['B3']='Start Date'
            ws1['C3']='Jobname'
            ws1['D3']='Task'
            ws1['E3']='Reference'
            ws1['F3']='Status'
            ws1['G3']='Finish Date'
            ws1['H3']='Remark'
            ws1['B1']=emp.name
            
            rows=4
            cols=2
            i=0
            # loop and store data in excel
            for job in jobs:
                
                if job.id in jsj:
                    
                    ws1.cell(rows+i,cols).value=job.sdate
                    ws1.cell(rows+i,cols+1).value=job.jobname
                    ws1.cell(rows+i,cols+2).value=job.task
                    ws1.cell(rows+i,cols+3).value=job.ref
                    ws1.cell(rows+i,cols+4).value=job.status
                    ws1.cell(rows+i,cols+5).value=job.fdate
                    ws1.cell(rows+i,cols+6).value=job.remark

                i=i+1
            # save as report.xlsx in base dir
            wb.save('report.xlsx')
            emp = Employee.objects.get(id=uid)
            emp.tempt = lenj
            emp.tempp = nub
            emp.tempc = comp
            emp.save()   
            context = {'emps': emps,'jobs':jobs,'jsj': jsj,'jsjn':jsjn,'jsjc':jsjc,'uid':uid}
            return render(request,'rep.html',context)
        # for report of all employee
        except:
            if fro == '':
                jobs = Job.objects.get(id=1)
                fro = jobs.day
                fro = int(fro.replace("-", ""))
            else:
                fro = int(fro.replace("-", ""))
            if to == '':
                now = datetime.datetime.now()
                sdate = f'{now.year}{now.month}{now.day}'
                to = int(sdate.replace("-", ""))
            else:
                to = int(to.replace("-", ""))
            # get number of jobs
            for j in emps:
                lenj = 0
                nub = 0
                bewu = 0
                for i in range(fro, to):
                    jobs = Job.objects.filter(day=i, user_id=j.id)  # total
                    lenj = lenj+len(jobs)
                    jobss = Job.objects.filter(
                        user_id=j.id, day=i, status='Not completed')  # not completed
                    nub = nub+len(jobss)
                    comp = lenj-nub
                # refresh employee status
                emp = Employee.objects.get(id=j.id)
                emp.tempt = lenj
                emp.tempp = nub
                emp.tempc = comp
                emp.save()

        return redirect('rep')
    context = {'emps': emps}
    return render(request, 'rep.html', context)

# list employee
@login_required(login_url='login')
def listemp(request):
    emps = Employee.objects.all()
    context = {'emps': emps, }
    return render(request, 'listemp.html', context)

# Delete employee from table
@login_required(login_url='login')
def delb(request, id):
    bks = Employee.objects.get(id=id)
    nsj = Job.objects.filter(user_id=id)

    nsj.delete()
    bks.delete()
    messages.success(request, "Employee removed successfully")
    return redirect("list")

# edit employee table
@login_required(login_url='login')
def edit(request, pk):
    bks = Employee.objects.get(id=pk)
    if request.method == 'POST':
        data = request.POST
        emp = data['empname']
        detail = data['details']
        phone = data['phone']
        email = data['email']
        bks.name = emp
        bks.email = email
        bks.detail = detail
        bks.phone = phone
        bks.save()
        messages.success(request, "Employee details updated")
    data = {'bks': bks}
    return render(request, 'editemployee.html', data)


# download xlsx file from base dir
def download_file(request,filename=''):
    if filename != '':
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # Define the full file path
        filepath = BASE_DIR + '/' + filename
        # Open the file for reading content
        path = open(filepath, 'rb')
        # Set the mime type
        mime_type, _ = mimetypes.guess_type(filepath)
        # Set the return value of the HttpResponse
        response = HttpResponse(path, content_type=mime_type)
        # Set the HTTP header for sending to browser
        response['Content-Disposition'] = "attachment; filename=%s" % filename
        # Return the response value
        return response
    else:
        # Load the template
        return render(request, 'reo.html')